"""
Exporters pour DataTable - Export Excel et CSV

Ce module fournit des classes pour exporter les données DataTable vers différents formats :
- Excel (.xlsx) avec openpyxl
- CSV (.csv) avec le module csv de Python

ARCHITECTURE:
- Interface IDataTableExporter pour l'extensibilité
- ExcelExporter pour les exports Excel
- CSVExporter pour les exports CSV
- ExportMixin pour intégration facile dans les vues

UTILISATION:
    class MyView(ServerSideDataTableView):
        enable_export = True  # Activé par défaut
        export_formats = ['excel', 'csv']  # Formats supportés
        export_filename = 'export'  # Nom du fichier (sans extension)

PRINCIPES SOLID:
- Single Responsibility: Chaque exporter gère un format
- Open/Closed: Extensible via IDataTableExporter
- Interface Segregation: Interface simple et claire
"""

from abc import ABC, abstractmethod
from typing import List, Dict, Any, Optional
from django.http import HttpResponse
from django.db.models import QuerySet
from rest_framework.serializers import Serializer
import logging
import csv
from datetime import datetime, date
from decimal import Decimal
import json

logger = logging.getLogger(__name__)


# =============================================================================
# UTILITAIRES POUR TRAITEMENT DES OBJETS IMBRIQUÉS
# =============================================================================

def flatten_nested_objects(data: List[Dict]) -> List[Dict]:
    """
    Aplatit les objets imbriqués et JSON en colonnes séparées
    
    Args:
        data: Liste de dictionnaires à traiter
        
    Returns:
        Liste de dictionnaires avec objets aplatis
        
    Exemple:
        Input: [{"id": 1, "details": {"produit": "MOB-3ans", "prix": 100}}]
        Output: [{"id": 1, "details_produit": "MOB-3ans", "details_prix": 100}]
        
    Version améliorée avec gestion d'erreurs robuste
    """
    if not data:
        return data
    
    flattened_data = []
    
    for row_index, row in enumerate(data):
        flattened_row = {}
        
        for key, value in row.items():
            try:
                # Traiter les objets imbriqués
                if isinstance(value, dict):
                    # Aplatir le dictionnaire
                    try:
                        for nested_key, nested_value in value.items():
                            new_key = f"{key}_{nested_key}"
                            # Éviter les clés avec des caractères invalides
                            new_key = new_key.replace(' ', '_').replace('/', '_').replace('\\', '_')
                            flattened_row[new_key] = nested_value
                    except Exception as e:
                        logger.warning(f"Erreur aplatissement dict {key}: {str(e)}")
                        flattened_row[key] = str(value)[:100]
                        
                elif isinstance(value, list):
                    # Traiter les listes
                    try:
                        if len(value) > 0 and isinstance(value[0], dict):
                            # Liste de dictionnaires - aplatir seulement le premier
                            for nested_key, nested_value in value[0].items():
                                new_key = f"{key}_0_{nested_key}"
                                new_key = new_key.replace(' ', '_').replace('/', '_').replace('\\', '_')
                                flattened_row[new_key] = nested_value
                        else:
                            # Liste simple - convertir en string
                            if len(value) <= 5:
                                flattened_row[key] = ", ".join(str(v) for v in value if v is not None)
                            else:
                                flattened_row[key] = f"{len(value)} éléments"
                    except Exception as e:
                        logger.warning(f"Erreur aplatissement list {key}: {str(e)}")
                        flattened_row[key] = f"{len(value)} éléments"
                            
                elif _is_json_string(value):
                    # Traiter les strings JSON
                    try:
                        json_data = json.loads(value)
                        if isinstance(json_data, dict):
                            for json_key, json_value in json_data.items():
                                new_key = f"{key}_{json_key}"
                                new_key = new_key.replace(' ', '_').replace('/', '_').replace('\\', '_')
                                flattened_row[new_key] = json_value
                        elif isinstance(json_data, list):
                            if len(json_data) <= 5:
                                flattened_row[key] = ", ".join(str(v) for v in json_data)
                            else:
                                flattened_row[key] = f"{len(json_data)} éléments"
                        else:
                            flattened_row[key] = value
                    except (json.JSONDecodeError, TypeError) as e:
                        logger.debug(f"String JSON invalide pour {key}: {str(e)}")
                        flattened_row[key] = value
                        
                else:
                    # Valeur simple - garder telle quelle
                    flattened_row[key] = value
                    
            except Exception as e:
                # En cas d'erreur sur cette clé, garder la valeur originale en string
                logger.warning(f"Erreur traitement ligne {row_index}, clé {key}: {str(e)}")
                try:
                    flattened_row[key] = str(value)[:100] if value is not None else ""
                except:
                    flattened_row[key] = ""
        
        flattened_data.append(flattened_row)
    
    return flattened_data


def _is_json_string(value) -> bool:
    """
    Détermine si une valeur est une chaîne JSON valide
    
    Args:
        value: Valeur à tester
        
    Returns:
        True si la valeur ressemble à du JSON
    """
    if not isinstance(value, str) or not value.strip():
        return False
    
    value = value.strip()
    return (value.startswith('{') and value.endswith('}')) or \
           (value.startswith('[') and value.endswith(']'))


# Import conditionnel pour Excel (openpyxl)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl n'est pas installé. L'export Excel ne sera pas disponible.")


# =============================================================================
# INTERFACE EXPORTER
# =============================================================================

class IDataTableExporter(ABC):
    """
    Interface pour les exporters DataTable
    
    Cette interface définit le contrat pour tous les exporters de données.
    """
    
    @abstractmethod
    def export(self, queryset: QuerySet, serializer_class: Optional[type] = None, 
               filename: str = 'export') -> HttpResponse:
        """
        Exporte les données du queryset
        
        Args:
            queryset: QuerySet à exporter
            serializer_class: Classe de serializer pour formater les données
            filename: Nom du fichier (sans extension)
            
        Returns:
            HttpResponse avec le fichier à télécharger
        """
        pass


# =============================================================================
# EXCEL EXPORTER
# =============================================================================

class ExcelExporter(IDataTableExporter):
    """
    Exporter pour fichiers Excel (.xlsx)
    
    Utilise openpyxl pour générer des fichiers Excel avec :
    - En-têtes formatés (gras, fond gris)
    - Colonnes auto-dimensionnées
    - Formatage des dates et nombres
    - Support des relations imbriquées
    """
    
    def __init__(self, 
                 sheet_name: str = 'Data',
                 header_style: bool = True,
                 auto_width: bool = True,
                 flatten_nested: bool = True):
        """
        Initialise l'exporter Excel
        
        Args:
            sheet_name: Nom de la feuille Excel
            header_style: Appliquer un style aux en-têtes
            auto_width: Ajuster automatiquement la largeur des colonnes
            flatten_nested: Aplatir les objets imbriqués en colonnes séparées
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl est requis pour l'export Excel. "
                "Installez-le avec: pip install openpyxl"
            )
        
        self.sheet_name = sheet_name
        self.header_style = header_style
        self.auto_width = auto_width
        self.flatten_nested = flatten_nested
    
    def export(self, queryset: QuerySet, serializer_class: Optional[type] = None, 
               filename: str = 'export') -> HttpResponse:
        """
        Exporte vers Excel
        
        Args:
            queryset: QuerySet à exporter
            serializer_class: Classe de serializer DRF
            filename: Nom du fichier sans extension
            
        Returns:
            HttpResponse avec fichier Excel
            
        Version améliorée avec gestion d'erreurs robuste
        """
        try:
            logger.info(f"Début export Excel: {filename}.xlsx - {queryset.count()} items")
            
            # Créer le workbook
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            
            # Sérialiser les données avec gestion d'erreurs
            try:
                if serializer_class:
                    logger.debug(f"Sérialisation avec {serializer_class.__name__}")
                    serializer = serializer_class(queryset, many=True)
                    data = serializer.data
                    # Convertir OrderedDict en dict Python standard
                    data = [dict(item) for item in data]
                else:
                    logger.debug("Sérialisation avec values()")
                    data = list(queryset.values())
            except Exception as e:
                logger.error(f"Erreur sérialisation: {str(e)}", exc_info=True)
                # Fallback: utiliser values() simple
                data = list(queryset.values())
            
            if not data:
                logger.warning("Aucune donnée à exporter")
                # Créer un fichier vide avec message
                ws.append(["Aucune donnée à exporter"])
            else:
                # Aplatir les objets imbriqués si activé
                if self.flatten_nested:
                    try:
                        logger.debug(f"Aplatissement de {len(data)} lignes...")
                        data = flatten_nested_objects(data)
                        logger.info(f"Objets imbriqués aplatis: {len(data)} lignes traitées")
                    except Exception as e:
                        logger.error(f"Erreur aplatissement: {str(e)}", exc_info=True)
                        # Continuer sans aplatissement
                        pass
                
                # Extraire les en-têtes
                try:
                    headers = list(data[0].keys())
                    # Nettoyer les noms de colonnes
                    headers = [str(h).replace('\x00', '').replace('\r', '').replace('\n', ' ')[:255] for h in headers]
                except Exception as e:
                    logger.error(f"Erreur extraction en-têtes: {str(e)}")
                    headers = ['data']
                
                # Écrire les en-têtes
                try:
                    ws.append(headers)
                    
                    # Appliquer le style aux en-têtes
                    if self.header_style:
                        self._apply_header_style(ws, len(headers))
                except Exception as e:
                    logger.error(f"Erreur écriture en-têtes: {str(e)}")
                
                # Écrire les données ligne par ligne avec gestion d'erreurs
                rows_written = 0
                for row_index, row_data in enumerate(data):
                    try:
                        row = [self._format_value(row_data.get(header)) for header in headers]
                        ws.append(row)
                        rows_written += 1
                    except Exception as e:
                        logger.warning(f"Erreur écriture ligne {row_index}: {str(e)}")
                        # Écrire une ligne avec des valeurs par défaut
                        try:
                            ws.append(["Erreur"] * len(headers))
                        except:
                            pass
                
                logger.info(f"{rows_written}/{len(data)} lignes écrites")
                
                # Ajuster la largeur des colonnes
                if self.auto_width:
                    try:
                        self._auto_size_columns(ws, headers, data)
                    except Exception as e:
                        logger.warning(f"Erreur auto-width: {str(e)}")
            
            # Créer la réponse HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            
            # Sauvegarder le workbook dans la réponse
            try:
                wb.save(response)
                logger.info(f"Export Excel réussi: {filename}.xlsx sauvegardé")
            except Exception as e:
                logger.error(f"Erreur sauvegarde Excel: {str(e)}", exc_info=True)
                raise
            
            return response
            
        except Exception as e:
            logger.error(f"Erreur CRITIQUE lors de l'export Excel: {str(e)}", exc_info=True)
            # Ne pas lever l'exception, retourner un fichier Excel avec message d'erreur
            try:
                wb = Workbook()
                ws = wb.active
                ws.append(["Erreur lors de l'export"])
                ws.append([str(e)])
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}_error.xlsx"'
                wb.save(response)
                return response
            except:
                # Dernier recours
                raise
    
    def _apply_header_style(self, ws, num_columns: int):
        """Applique un style aux en-têtes"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _auto_size_columns(self, ws, headers: List[str], data: List[Dict]):
        """Ajuste automatiquement la largeur des colonnes"""
        for idx, header in enumerate(headers, 1):
            # Largeur basée sur le header
            max_length = len(str(header))
            
            # Vérifier les 100 premières lignes pour optimiser
            for row_data in data[:100]:
                value = str(row_data.get(header, ''))
                max_length = max(max_length, len(value))
            
            # Limiter la largeur max à 50 caractères
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = adjusted_width
    
    def _format_value(self, value: Any) -> Any:
        """
        Formate une valeur pour Excel
        
        Convertit les types Python en types compatibles Excel
        Version améliorée avec gestion d'erreurs robuste
        """
        try:
            if value is None:
                return ""
            elif isinstance(value, (datetime, date)):
                try:
                    return value.strftime("%Y-%m-%d %H:%M:%S") if isinstance(value, datetime) else value.strftime("%Y-%m-%d")
                except:
                    return str(value)
            elif isinstance(value, Decimal):
                try:
                    return float(value)
                except:
                    return str(value)
            elif isinstance(value, bool):
                return "Oui" if value else "Non"
            elif isinstance(value, (list, dict)):
                # Si on arrive ici, c'est que l'aplatissement n'a pas été fait
                # On affiche une représentation lisible
                try:
                    if isinstance(value, dict):
                        # Filtrer les valeurs None et limiter la longueur
                        items = [f"{k}: {v}" for k, v in value.items() if v is not None]
                        return ", ".join(items[:10])  # Limite à 10 éléments
                    elif isinstance(value, list):
                        items = [str(v) for v in value if v is not None]
                        return ", ".join(items[:10])  # Limite à 10 éléments
                except:
                    return str(value)[:100]  # Limite à 100 caractères
            else:
                # Valeur simple - convertir en string safe
                str_value = str(value)
                # Nettoyer les caractères problématiques pour Excel
                return str_value.replace('\x00', '').replace('\r', '').replace('\n', ' ')[:1000]
        except Exception as e:
            # Dernier recours : retourner string vide ou valeur brute
            logger.warning(f"Erreur formatage valeur {type(value).__name__}: {str(e)}")
            try:
                return str(value)[:100] if value is not None else ""
            except:
                return ""


# =============================================================================
# CSV EXPORTER
# =============================================================================

class CSVExporter(IDataTableExporter):
    """
    Exporter pour fichiers CSV
    
    Utilise le module csv de Python pour générer des fichiers CSV avec :
    - Encodage UTF-8 avec BOM pour Excel
    - Délimiteur configurable
    - Échappement automatique des guillemets
    """
    
    def __init__(self, 
                 delimiter: str = ',',
                 encoding: str = 'utf-8-sig',  # utf-8-sig pour Excel
                 flatten_nested: bool = True):
        """
        Initialise l'exporter CSV
        
        Args:
            delimiter: Délimiteur CSV (virgule par défaut)
            encoding: Encodage du fichier (utf-8-sig pour Excel)
            flatten_nested: Aplatir les objets imbriqués en colonnes séparées
        """
        self.delimiter = delimiter
        self.encoding = encoding
        self.flatten_nested = flatten_nested
    
    def export(self, queryset: QuerySet, serializer_class: Optional[type] = None, 
               filename: str = 'export') -> HttpResponse:
        """
        Exporte vers CSV
        
        Args:
            queryset: QuerySet à exporter
            serializer_class: Classe de serializer DRF
            filename: Nom du fichier sans extension
            
        Returns:
            HttpResponse avec fichier CSV
            
        Version améliorée avec gestion d'erreurs robuste
        """
        try:
            logger.info(f"Début export CSV: {filename}.csv - {queryset.count()} items")
            
            # Sérialiser les données avec gestion d'erreurs
            try:
                if serializer_class:
                    logger.debug(f"Sérialisation avec {serializer_class.__name__}")
                    serializer = serializer_class(queryset, many=True)
                    data = serializer.data
                    # Convertir OrderedDict en dict Python standard
                    data = [dict(item) for item in data]
                else:
                    logger.debug("Sérialisation avec values()")
                    data = list(queryset.values())
            except Exception as e:
                logger.error(f"Erreur sérialisation: {str(e)}", exc_info=True)
                # Fallback: utiliser values() simple
                data = list(queryset.values())
            
            # Aplatir les objets imbriqués si activé
            if self.flatten_nested:
                try:
                    logger.debug(f"Aplatissement de {len(data)} lignes...")
                    data = flatten_nested_objects(data)
                    logger.info(f"Objets imbriqués aplatis: {len(data)} lignes traitées")
                except Exception as e:
                    logger.error(f"Erreur aplatissement: {str(e)}", exc_info=True)
                    # Continuer sans aplatissement
                    pass
            
            # Créer la réponse HTTP
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            
            if not data:
                logger.warning("Aucune donnée à exporter")
                writer = csv.writer(response, delimiter=self.delimiter)
                writer.writerow(["Aucune donnée à exporter"])
            else:
                try:
                    # Extraire les en-têtes
                    headers = list(data[0].keys())
                    # Nettoyer les noms de colonnes
                    headers = [str(h).replace('\x00', '').replace('\r', '').replace('\n', ' ')[:255] for h in headers]
                    
                    # Créer le writer CSV
                    writer = csv.DictWriter(response, fieldnames=headers, delimiter=self.delimiter, 
                                          extrasaction='ignore')  # Ignore les clés supplémentaires
                    
                    # Écrire les en-têtes
                    writer.writeheader()
                    
                    # Écrire les données ligne par ligne
                    rows_written = 0
                    for row_index, row_data in enumerate(data):
                        try:
                            # Formater les valeurs
                            formatted_row = {
                                key: self._format_value(value) 
                                for key, value in row_data.items()
                                if key in headers  # Seulement les colonnes valides
                            }
                            writer.writerow(formatted_row)
                            rows_written += 1
                        except Exception as e:
                            logger.warning(f"Erreur écriture ligne CSV {row_index}: {str(e)}")
                            # Écrire une ligne avec des valeurs par défaut
                            try:
                                writer.writerow({h: "Erreur" for h in headers})
                            except:
                                pass
                    
                    logger.info(f"{rows_written}/{len(data)} lignes CSV écrites")
                    
                except Exception as e:
                    logger.error(f"Erreur structure CSV: {str(e)}", exc_info=True)
                    # Écrire au moins un message
                    writer = csv.writer(response, delimiter=self.delimiter)
                    writer.writerow(["Erreur lors de la génération du CSV"])
                    writer.writerow([str(e)])
            
            logger.info(f"Export CSV réussi: {filename}.csv")
            return response
            
        except Exception as e:
            logger.error(f"Erreur CRITIQUE lors de l'export CSV: {str(e)}", exc_info=True)
            # Ne pas lever l'exception, retourner un CSV avec message d'erreur
            try:
                response = HttpResponse(content_type='text/csv; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="{filename}_error.csv"'
                writer = csv.writer(response, delimiter=self.delimiter)
                writer.writerow(["Erreur lors de l'export"])
                writer.writerow([str(e)])
                return response
            except:
                # Dernier recours
                raise
    
    def _format_value(self, value: Any) -> str:
        """
        Formate une valeur pour CSV
        
        Convertit toutes les valeurs en chaînes de caractères
        """
        if value is None:
            return ""
        elif isinstance(value, (datetime, date)):
            return value.strftime("%Y-%m-%d %H:%M:%S") if isinstance(value, datetime) else value.strftime("%Y-%m-%d")
        elif isinstance(value, bool):
            return "Oui" if value else "Non"
        elif isinstance(value, (list, dict)):
            return str(value)
        else:
            return str(value)


# =============================================================================
# EXPORT MANAGER
# =============================================================================

class ExportManager:
    """
    Gestionnaire d'exports DataTable
    
    Centralise la gestion des différents formats d'export :
    - Enregistrement des exporters
    - Détection du format demandé
    - Dispatch vers l'exporter approprié
    """
    
    def __init__(self):
        """Initialise le gestionnaire avec les exporters par défaut"""
        self._exporters: Dict[str, IDataTableExporter] = {}
        
        # Enregistrer les exporters par défaut avec aplatissement activé
        if OPENPYXL_AVAILABLE:
            self.register_exporter('excel', ExcelExporter(flatten_nested=True))
            self.register_exporter('xlsx', ExcelExporter(flatten_nested=True))
        
        self.register_exporter('csv', CSVExporter(flatten_nested=True))
    
    def register_exporter(self, format_name: str, exporter: IDataTableExporter):
        """
        Enregistre un exporter pour un format
        
        Args:
            format_name: Nom du format (ex: 'excel', 'csv', 'pdf')
            exporter: Instance de l'exporter
        """
        self._exporters[format_name.lower()] = exporter
        logger.debug(f"Exporter enregistré: {format_name}")
    
    def get_exporter(self, format_name: str) -> Optional[IDataTableExporter]:
        """
        Récupère un exporter par son nom
        
        Args:
            format_name: Nom du format
            
        Returns:
            Instance de l'exporter ou None
        """
        return self._exporters.get(format_name.lower())
    
    def is_format_supported(self, format_name: str) -> bool:
        """
        Vérifie si un format est supporté
        
        Args:
            format_name: Nom du format
            
        Returns:
            True si le format est supporté
        """
        return format_name.lower() in self._exporters
    
    def export(self, format_name: str, queryset: QuerySet, 
               serializer_class: Optional[type] = None,
               filename: str = 'export') -> HttpResponse:
        """
        Exporte vers le format demandé
        
        Args:
            format_name: Format d'export ('excel', 'csv', etc.)
            queryset: QuerySet à exporter
            serializer_class: Classe de serializer
            filename: Nom du fichier
            
        Returns:
            HttpResponse avec le fichier
            
        Raises:
            ValueError: Si le format n'est pas supporté
        """
        exporter = self.get_exporter(format_name)
        
        if not exporter:
            available_formats = ', '.join(self._exporters.keys())
            raise ValueError(
                f"Format d'export non supporté: {format_name}. "
                f"Formats disponibles: {available_formats}"
            )
        
        return exporter.export(queryset, serializer_class, filename)


# Instance globale du gestionnaire d'exports
export_manager = ExportManager()

