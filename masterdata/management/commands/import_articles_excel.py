"""
Commande Django pour importer des articles depuis un fichier Excel
Usage: python manage.py import_articles_excel <chemin_fichier.xlsx>
"""
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from masterdata.models import article, Compte, produit, marque, fournisseur, nature
from datetime import datetime


class Command(BaseCommand):
    help = 'Importe des articles depuis un fichier Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='Chemin vers le fichier Excel à importer'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Simuler l\'importation sans enregistrer dans la base de données'
        )
        parser.add_argument(
            '--skip-errors',
            action='store_true',
            help='Continuer l\'importation même en cas d\'erreurs sur certaines lignes'
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        dry_run = options['dry_run']
        skip_errors = options['skip_errors']

        if dry_run:
            self.stdout.write(self.style.WARNING('🔍 Mode DRY-RUN activé - Aucune modification ne sera enregistrée'))

        try:
            # Charger le fichier Excel
            self.stdout.write(f'📂 Chargement du fichier: {excel_file}')
            wb = openpyxl.load_workbook(excel_file)
            
            # Lire la feuille "Articles à importer"
            if "Articles à importer" not in wb.sheetnames:
                raise CommandError('❌ La feuille "Articles à importer" est introuvable dans le fichier Excel')
            
            ws = wb["Articles à importer"]
            self.stdout.write(self.style.SUCCESS('✅ Fichier chargé avec succès'))

            # Lire les en-têtes (première ligne)
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            # Vérifier les colonnes obligatoires
            required_columns = ['designation', 'date_achat', 'prix_achat', 'qte', 'N_facture', 'compte_numero', 'produit_code']
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                raise CommandError(f'❌ Colonnes obligatoires manquantes: {", ".join(missing_columns)}')

            # Statistiques
            stats = {
                'total': 0,
                'success': 0,
                'errors': 0,
                'skipped': 0
            }

            errors_list = []

            # Traiter chaque ligne (à partir de la ligne 2)
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    stats['total'] += 1
                    
                    # Créer un dictionnaire des valeurs de la ligne
                    row_data = dict(zip(headers, row))
                    
                    # Ignorer les lignes vides (si designation est vide)
                    if not row_data.get('designation'):
                        stats['skipped'] += 1
                        self.stdout.write(self.style.WARNING(f'⏭️  Ligne {row_num}: Ignorée (designation vide)'))
                        continue

                    try:
                        # Préparer les données de l'article
                        article_data = self._prepare_article_data(row_data, row_num)
                        
                        if not dry_run:
                            # Créer l'article
                            new_article = article.objects.create(**article_data)
                            stats['success'] += 1
                            self.stdout.write(self.style.SUCCESS(
                                f'✅ Ligne {row_num}: Article créé - {new_article.designation} '
                                f'(Code: {new_article.code_article})'
                            ))
                        else:
                            stats['success'] += 1
                            self.stdout.write(self.style.SUCCESS(
                                f'✅ Ligne {row_num}: Article valide - {row_data["designation"]} '
                                f'(simulation)'
                            ))

                    except Exception as e:
                        stats['errors'] += 1
                        error_msg = f'❌ Ligne {row_num}: {str(e)}'
                        errors_list.append(error_msg)
                        self.stderr.write(self.style.ERROR(error_msg))
                        
                        if not skip_errors:
                            raise CommandError(f'Importation interrompue à la ligne {row_num}. Utilisez --skip-errors pour continuer malgré les erreurs.')

                # Si dry-run, annuler la transaction
                if dry_run:
                    transaction.set_rollback(True)
                    self.stdout.write(self.style.WARNING('\n🔄 Transaction annulée (dry-run)'))

            # Afficher les statistiques
            self.stdout.write('\n' + '='*60)
            self.stdout.write(self.style.SUCCESS('📊 STATISTIQUES D\'IMPORTATION'))
            self.stdout.write('='*60)
            self.stdout.write(f'📝 Total de lignes traitées: {stats["total"]}')
            self.stdout.write(self.style.SUCCESS(f'✅ Succès: {stats["success"]}'))
            if stats['errors'] > 0:
                self.stdout.write(self.style.ERROR(f'❌ Erreurs: {stats["errors"]}'))
            if stats['skipped'] > 0:
                self.stdout.write(self.style.WARNING(f'⏭️  Ignorées: {stats["skipped"]}'))
            
            if errors_list:
                self.stdout.write('\n' + self.style.ERROR('📋 DÉTAIL DES ERREURS:'))
                for error in errors_list:
                    self.stdout.write(f'   {error}')

            if not dry_run and stats['success'] > 0:
                self.stdout.write('\n' + self.style.SUCCESS(f'🎉 {stats["success"]} article(s) importé(s) avec succès !'))
            elif dry_run:
                self.stdout.write('\n' + self.style.WARNING('ℹ️  Mode simulation - Aucune donnée n\'a été enregistrée'))

        except FileNotFoundError:
            raise CommandError(f'❌ Fichier introuvable: {excel_file}')
        except Exception as e:
            raise CommandError(f'❌ Erreur lors de l\'importation: {str(e)}')

    def _prepare_article_data(self, row_data, row_num):
        """Prépare et valide les données d'un article"""
        
        # Récupérer le compte
        compte_numero = row_data.get('compte_numero')
        if not compte_numero:
            raise ValueError('compte_numero est obligatoire')
        
        try:
            compte_obj = Compte.objects.get(numero=compte_numero)
        except Compte.DoesNotExist:
            raise ValueError(f'Compte avec numero "{compte_numero}" introuvable')

        # Récupérer le produit
        produit_code = row_data.get('produit_code')
        if not produit_code:
            raise ValueError('produit_code est obligatoire')
        
        try:
            produit_obj = produit.objects.get(code=produit_code)
        except produit.DoesNotExist:
            raise ValueError(f'Produit avec code "{produit_code}" introuvable')

        # Récupérer la marque (optionnel)
        marque_obj = None
        marque_nom = row_data.get('marque_nom')
        if marque_nom:
            try:
                marque_obj = marque.objects.get(nom=marque_nom)
            except marque.DoesNotExist:
                raise ValueError(f'Marque avec nom "{marque_nom}" introuvable')

        # Récupérer le fournisseur (optionnel)
        fournisseur_obj = None
        fournisseur_nom = row_data.get('fournisseur_nom')
        if fournisseur_nom:
            try:
                fournisseur_obj = fournisseur.objects.get(nom=fournisseur_nom)
            except fournisseur.DoesNotExist:
                raise ValueError(f'Fournisseur avec nom "{fournisseur_nom}" introuvable')

        # Récupérer la nature (optionnel)
        nature_obj = None
        nature_code = row_data.get('nature_code')
        if nature_code:
            try:
                nature_obj = nature.objects.get(code=nature_code)
            except nature.DoesNotExist:
                raise ValueError(f'Nature avec code "{nature_code}" introuvable')

        # Parser les dates
        date_achat = self._parse_date(row_data.get('date_achat'), 'date_achat')
        date_expiration = self._parse_date(row_data.get('date_expiration'), 'date_expiration', required=False)
        date_peremption = self._parse_date(row_data.get('date_peremption'), 'date_peremption', required=False)

        # Valider valider (booléen)
        valider_value = True
        if row_data.get('valider'):
            valider_str = str(row_data.get('valider')).upper()
            valider_value = valider_str in ['TRUE', '1', 'OUI', 'YES']

        # Construire les données de l'article
        article_data = {
            'code_article': row_data.get('code_article') or '',  # Auto-généré si vide
            'designation': row_data.get('designation'),
            'date_achat': date_achat,
            'numero_comptable': row_data.get('numero_comptable') or '',
            'couleur': row_data.get('couleur') or '',
            'poids': row_data.get('poids'),
            'volume': row_data.get('volume'),
            'langueur': row_data.get('langueur'),
            'hauteur': row_data.get('hauteur'),
            'largeur': row_data.get('largeur'),
            'date_expiration': date_expiration,
            'date_peremption': date_peremption,
            'prix_achat': float(row_data.get('prix_achat', 0)),
            'qte': int(row_data.get('qte', 1)),
            'qte_recue': int(row_data.get('qte_recue')) if row_data.get('qte_recue') else None,
            'N_facture': row_data.get('N_facture'),
            'valider': valider_value,
            'compte': compte_obj,
            'produit': produit_obj,
            'marque': marque_obj,
            'fournisseur': fournisseur_obj,
            'nature': nature_obj,
        }

        return article_data

    def _parse_date(self, date_value, field_name, required=True):
        """Parse une date depuis Excel"""
        if not date_value:
            if required:
                raise ValueError(f'{field_name} est obligatoire')
            return None

        # Si c'est déjà un objet datetime (depuis Excel)
        if isinstance(date_value, datetime):
            return date_value.date()

        # Si c'est une chaîne, essayer de la parser
        if isinstance(date_value, str):
            try:
                # Format YYYY-MM-DD
                return datetime.strptime(date_value, '%Y-%m-%d').date()
            except ValueError:
                try:
                    # Format DD/MM/YYYY
                    return datetime.strptime(date_value, '%d/%m/%Y').date()
                except ValueError:
                    raise ValueError(f'{field_name}: Format de date invalide (attendu: YYYY-MM-DD ou DD/MM/YYYY)')

        raise ValueError(f'{field_name}: Type de date invalide')

