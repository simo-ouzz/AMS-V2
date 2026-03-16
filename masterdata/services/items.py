from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Optional

from django.core.files.uploadedfile import UploadedFile
from django.db import transaction
from openpyxl import load_workbook

from masterdata.models import ArchiveItem, item


class ArchiveItemsImportError(Exception):
    """
    Erreur fonctionnelle pour l'import Excel des archivages d'items.
    Contient la liste complète des messages d'erreur détectés.
    """

    def __init__(self, errors: List[str]):
        self.errors = errors
        super().__init__("; ".join(errors))
from masterdata.repositories.item_repository import (
    archive_item_instance,
    bulk_archive_items_from_rows,
    get_item_by_id,
    list_archive_items_by_compte,
)


@dataclass
class ArchiveItemResult:
    success: bool
    message: str
    already_archived: bool = False


def archive_or_unarchive_item_service(
    *, item_id: int, action: str, commentaire: str | None = None
) -> ArchiveItemResult:
    """
    Logique métier d'archivage / désarchivage d'un item unique.
    """
    it = get_item_by_id(item_id)

    if action == "archive":
        if it.archive:
            return ArchiveItemResult(
                success=True, already_archived=True, message="Item déjà archivé."
            )
        archive_item_instance(it, commentaire)
        return ArchiveItemResult(
            success=True, already_archived=False, message="Item archivé avec succès."
        )

    if action == "unarchive":
        if not it.archive:
            return ArchiveItemResult(
                success=True,
                already_archived=False,
                message="Item déjà désarchivé.",
            )
        from masterdata.repositories.item_repository import unarchive_item_instance

        unarchive_item_instance(it)
        return ArchiveItemResult(success=True, message="Item désarchivé avec succès.")

    raise ValueError('Action invalide. Utiliser "archive" ou "unarchive".')


def list_archive_items_for_user(user) -> List[ArchiveItem]:
    """
    Retourne la liste des ArchiveItem pour le compte de l'utilisateur.
    """
    if not getattr(user, "compte", None):
        return []
    return list(list_archive_items_by_compte(user.compte))


def _iter_rows_from_workbook(wb) -> List[Dict[str, Optional[str]]]:
    """
    Extrait les lignes (tag, designation, commentaire) depuis le premier onglet du classeur.

    Colonnes attendues dans la première ligne (en-têtes) :
    - obligatoire : 'tag'
    - optionnelles : 'designation', 'commentaire'
    """
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    header_index: Dict[str, int] = {
        str(h).strip(): idx for idx, h in enumerate(headers) if h
    }
    if "tag" not in header_index:
        raise ValueError("La colonne 'tag' est obligatoire dans le fichier Excel.")

    tag_idx = header_index["tag"]
    designation_idx = header_index.get("designation")
    commentaire_idx = header_index.get("commentaire")

    rows: List[Dict[str, Optional[str]]] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_tag = row[tag_idx]
        if raw_tag in (None, ""):
            continue

        tag_ref = str(raw_tag).strip()

        designation: Optional[str] = None
        if designation_idx is not None and designation_idx < len(row):
            designation_val = row[designation_idx]
            if designation_val not in (None, ""):
                designation = str(designation_val).strip()

        commentaire: Optional[str] = None
        if commentaire_idx is not None and commentaire_idx < len(row):
            commentaire_val = row[commentaire_idx]
            if commentaire_val not in (None, ""):
                commentaire = str(commentaire_val)

        rows.append(
            {
                "row": idx,
                "tag": tag_ref,
                "designation": designation,
                "commentaire": commentaire,
            }
        )

    return rows


def import_archive_items_from_excel_service(
    *, user, excel_file: UploadedFile, skip_errors: bool = True
) -> Dict[str, Any]:
    """
    Importe un fichier Excel contenant des lignes (tag, designation, commentaire) pour archiver des items.

    Règles :
    - Si l'item n'existe pas : considéré comme erreur.
    - Si l'item est déjà archivé : compté dans 'already_archived', pas réarchivé.
    - Sinon : archivé et historisé.
    """
    if not getattr(user, "is_authenticated", False):
        raise PermissionError("Utilisateur non authentifié.")

    try:
        wb = load_workbook(excel_file)
        rows = _iter_rows_from_workbook(wb)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Erreur lors de la lecture du fichier Excel : {exc}") from exc

    total = 0
    success = 0
    already_archived = 0
    not_found = 0

    errors: List[str] = []

    with transaction.atomic():
        for row in rows:
            total += 1
            excel_row = row.get("row")
            tag_ref = row["tag"]
            commentaire = row["commentaire"]
            expected_designation = row["designation"]

            try:
                it = item.objects.select_related("article", "tag").get(
                    tag__reference=tag_ref,
                    article__compte=user.compte,
                )
            except item.DoesNotExist:
                not_found += 1
                msg = (
                    f"Ligne {excel_row}: Item avec tag='{tag_ref}' "
                    f"introuvable pour ce compte."
                )
                errors.append(msg)
                continue

            # Vérification facultative de la désignation si fournie
            if expected_designation and it.article and it.article.designation:
                if it.article.designation.strip() != expected_designation.strip():
                    errors.append(
                        f"Ligne {excel_row}: Incohérence de désignation pour tag='{tag_ref}' "
                        f"(fichier='{expected_designation}', base='{it.article.designation}')."
                    )

            if it.archive:
                already_archived += 1
                continue

            archive_item_instance(it, commentaire)
            success += 1

        # Tout ou rien : s'il y a des erreurs, on annule l'entièreté de la transaction
        if errors and not skip_errors:
            raise ArchiveItemsImportError(errors)

    return {
        "total": total,
        "success": success,
        "already_archived": already_archived,
        "not_found": not_found,
        "errors": errors,
    }


def archive_items_batch_service(
    *, user, items_id: List[int]
) -> Dict[str, Any]:
    """
    Archive une liste d'items en une seule opération.

    - `items_id` : liste d'IDs d'items
    """
    if not getattr(user, "is_authenticated", False):
        raise PermissionError("Utilisateur non authentifié.")

    total = len(items_id)
    success = 0
    already_archived = 0
    not_found = 0
    errors: List[str] = []

    for iid in items_id:
        try:
            result = archive_or_unarchive_item_service(
                item_id=int(iid), action="archive", commentaire=None
            )
            success += 1
            if result.already_archived:
                already_archived += 1
        except item.DoesNotExist:
            not_found += 1
            errors.append(f"Item avec id={iid} introuvable.")
        except ValueError as exc:
            errors.append(f"Item id={iid} : {exc}")
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Item id={iid} : {exc}")

    return {
        "total": total,
        "success": success,
        "already_archived": already_archived,
        "not_found": not_found,
        "errors": errors,
    }
